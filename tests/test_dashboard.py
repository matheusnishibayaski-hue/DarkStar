"""Testes do dashboard / histórico de scans."""

from __future__ import annotations

import tempfile
import unittest
import io
from pathlib import Path
from unittest.mock import patch

from backend.database import db as db_mod
from fastapi.testclient import TestClient

from tests.auth_patch import patch_chat_api_token


class TestDashboardStore(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        db_mod.reset_engine_for_tests()
        self._url = f"sqlite:///{(self.root / 't.db').as_posix()}"
        self.patches = [
            patch.object(db_mod, "DATABASE_URL", ""),
            patch.object(db_mod, "_SQLITE_PATH", self.root / "t.db"),
            patch.object(db_mod, "resolve_database_url", return_value=self._url),
        ]
        for p in self.patches:
            p.start()
        db_mod.reset_engine_for_tests()
        db_mod.init_db()

    def tearDown(self):
        for p in self.patches:
            p.stop()
        db_mod.reset_engine_for_tests()
        self.tmp.cleanup()

    def test_save_history_metrics(self):
        sid = "dash-sess-1"
        ok = db_mod.save_scan_result(
            {
                "scan_id": "abc123",
                "target": "lab.test",
                "risk_profile": "safe-active",
                "scan_profile": "basic",
                "vulnerability_count": 2,
                "critical": 1,
                "high": 1,
                "medium": 0,
                "low": 0,
                "chat_session_id": sid,
                "findings": [
                    {
                        "id": "1",
                        "title": "Missing HSTS",
                        "severity": "medium",
                        "remediation": "Enable HSTS",
                    },
                    {
                        "id": "2",
                        "title": "CVE-2020-1",
                        "severity": "critical",
                        "remediation": "Patch",
                    },
                ],
                "rounds": 3,
                "status": "completed",
                "scan_type": "test",
            }
        )
        self.assertTrue(ok)
        hist = db_mod.get_scan_history(days=30, session_id=sid)
        self.assertGreaterEqual(len(hist), 1)
        self.assertEqual(hist[0]["target"], "lab.test")
        metrics = db_mod.compute_metrics(days=30, session_id=sid)
        self.assertGreaterEqual(metrics["total_scans"], 1)
        tops = db_mod.get_top_issues(limit=5, session_id=sid)
        self.assertTrue(any(t["title"] for t in tops))
        trend = db_mod.vulnerability_trend(days=30, session_id=sid)
        self.assertTrue(isinstance(trend, list))
        bundle = db_mod.dashboard_bundle(days=30, session_id=sid)
        self.assertEqual(bundle["metrics"]["total_scans"], metrics["total_scans"])
        self.assertIn("trend", bundle)
        self.assertIn("history", bundle)
        # Outra conversa sem scans não herda opens globais
        empty_m = db_mod.compute_metrics(days=30, session_id="empty-other-sess")
        self.assertEqual(empty_m["total_scans"], 0)
        self.assertEqual(empty_m["open_vulnerabilities"], 0)
        self.assertEqual(db_mod.get_top_issues(limit=5, session_id="empty-other-sess"), [])


class TestDashboardRoutes(unittest.TestCase):
    def test_metrics_endpoint(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            url = f"sqlite:///{(root / 'r.db').as_posix()}"
            db_mod.reset_engine_for_tests()
            with (
                patch.object(db_mod, "DATABASE_URL", ""),
                patch.object(db_mod, "_SQLITE_PATH", root / "r.db"),
                patch.object(db_mod, "resolve_database_url", return_value=url),
                patch_chat_api_token(""),
            ):
                db_mod.reset_engine_for_tests()
                db_mod.init_db()
                sid = "dash-route-1"
                db_mod.save_scan_result(
                    {
                        "scan_id": "route1",
                        "target": "a.test",
                        "vulnerability_count": 0,
                        "critical": 0,
                        "high": 0,
                        "medium": 0,
                        "low": 0,
                        "findings": [],
                        "status": "completed",
                        "scan_type": "test",
                        "chat_session_id": sid,
                    }
                )
                from backend.main import app

                client = TestClient(app)
                q = f"days=30&session_id={sid}"
                r = client.get(f"/api/dashboard/metrics?{q}")
                self.assertEqual(r.status_code, 200)
                self.assertEqual(r.json()["status"], "ok")
                r2 = client.get(f"/api/dashboard/scan-history?{q}")
                self.assertEqual(r2.status_code, 200)
                bundle = client.get(f"/api/dashboard/bundle?{q}")
                self.assertEqual(bundle.status_code, 200)
                body = bundle.json()
                self.assertEqual(body["status"], "ok")
                self.assertIn("metrics", body)
                self.assertIn("trend", body)
                self.assertIn("top_issues", body)
                self.assertIn("history", body)
                r3 = client.get(f"/api/dashboard/export?format=json&{q}")
                self.assertEqual(r3.status_code, 200)
                xlsx = client.get(f"/api/dashboard/export?format=xlsx&{q}")
                self.assertEqual(xlsx.status_code, 200)
                self.assertIn(
                    "spreadsheetml",
                    xlsx.headers.get("content-type", ""),
                )
                self.assertTrue(xlsx.content[:2] == b"PK")
                from openpyxl import load_workbook

                wb = load_workbook(filename=io.BytesIO(xlsx.content))
                self.assertEqual(
                    set(wb.sheetnames),
                    {"Resumo", "Tendência", "Principais problemas", "Histórico de scans"},
                )
            db_mod.reset_engine_for_tests()


if __name__ == "__main__":
    unittest.main()
