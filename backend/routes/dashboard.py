"""API do dashboard — métricas, trends, histórico e export (por conversa)."""

from __future__ import annotations

import csv
import io
from datetime import datetime, timezone

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import Response, StreamingResponse

from backend.database.db import (
    compute_metrics,
    dashboard_bundle,
    get_scan_history,
    get_top_issues,
    purge_scans_for_session,
    summary_report,
    vulnerability_trend,
)

router = APIRouter(prefix="/api/dashboard", tags=["dashboard"])


def _require_session(session_id: str | None) -> str:
    sid = (session_id or "").strip()
    if not sid:
        raise HTTPException(
            status_code=400,
            detail="session_id required (dashboard is scoped per conversation)",
        )
    return sid


def _client_blurb(metrics: dict) -> str:
    scans = int(metrics.get("total_scans") or 0)
    open_v = int(metrics.get("open_vulnerabilities") or 0)
    if scans == 0:
        return "Ainda não há varreduras nesta conversa."
    if open_v > 0:
        return (
            f"{scans} varredura(s) no período. {open_v} problema(s) em aberto — "
            "priorize críticas e altas antes de entregar ao cliente."
        )
    return f"{scans} varredura(s) no período. Nenhum problema em aberto neste recorte."


def _build_xlsx(
    *,
    sid: str,
    days: int,
    metrics: dict,
    history: list,
    trend: list,
    top_issues: list,
) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl unavailable") from exc

    wb = Workbook()
    bold = Font(bold=True)

    ws = wb.active
    ws.title = "Resumo"
    ws.append(["Dashboard DarkStar — visão para cliente"])
    ws["A1"].font = bold
    ws.append(["Conversa", sid])
    ws.append(["Período (dias)", days])
    ws.append(["Gerado em", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")])
    ws.append([])
    ws.append(["Indicador", "Valor", "O que isso significa"])
    for cell in ws[6]:
        cell.font = bold
    rows_resumo = [
        (
            "Varreduras",
            metrics.get("total_scans", 0),
            "Quantidade de varreduras registradas no período.",
        ),
        (
            "Críticas (média)",
            round(float(metrics.get("avg_critical") or 0)),
            "Média de achados críticos por varredura.",
        ),
        (
            "Altas (média)",
            round(float(metrics.get("avg_high") or 0)),
            "Média de achados altos por varredura.",
        ),
        (
            "Problemas em aberto",
            metrics.get("open_vulnerabilities", 0),
            "Itens ainda sem fechamento nesta conversa.",
        ),
        (
            "Total de achados (soma)",
            metrics.get("total_vulnerabilities", 0),
            "Soma de problemas encontrados nas varreduras.",
        ),
    ]
    for row in rows_resumo:
        ws.append(list(row))
    ws.append([])
    ws.append(["Texto para o cliente", _client_blurb(metrics)])

    ws_t = wb.create_sheet("Tendência")
    ws_t.append(["Data", "Críticas", "Altas", "Médias", "Baixas", "Total"])
    for cell in ws_t[1]:
        cell.font = bold
    for r in trend:
        ws_t.append(
            [
                r.get("date"),
                r.get("critical") or 0,
                r.get("high") or 0,
                r.get("medium") or 0,
                r.get("low") or 0,
                r.get("total") or 0,
            ]
        )
    if not trend:
        ws_t.append(["—", 0, 0, 0, 0, 0])

    ws_i = wb.create_sheet("Principais problemas")
    ws_i.append(["Problema", "Gravidade", "Alvo", "Vezes"])
    for cell in ws_i[1]:
        cell.font = bold
    for issue in top_issues:
        ws_i.append(
            [
                issue.get("title") or "",
                issue.get("severity") or "",
                issue.get("target") or "",
                issue.get("count") or 0,
            ]
        )
    if not top_issues:
        ws_i.append(["Nenhum problema nesta conversa", "", "", 0])

    ws_h = wb.create_sheet("Histórico de scans")
    ws_h.append(
        [
            "Alvo",
            "Problemas",
            "Críticas",
            "Altas",
            "Médias",
            "Baixas",
            "Status",
            "Tipo",
            "Quando",
        ]
    )
    for cell in ws_h[1]:
        cell.font = bold
    for row in history:
        when = str(row.get("timestamp") or "")[:19].replace("T", " ")
        ws_h.append(
            [
                row.get("target") or "",
                row.get("vulnerability_count") or 0,
                row.get("critical") or 0,
                row.get("high") or 0,
                row.get("medium") or 0,
                row.get("low") or 0,
                row.get("status") or "",
                row.get("scan_type") or "",
                when,
            ]
        )
    if not history:
        ws_h.append(["Nenhuma varredura", 0, 0, 0, 0, 0, "", "", ""])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@router.get("/metrics")
def api_metrics(
    days: int = Query(30, ge=1, le=365),
    session_id: str = Query(..., min_length=1, max_length=128),
):
    sid = _require_session(session_id)
    data = compute_metrics(days=days, session_id=sid)
    return {
        "status": "ok",
        "data": data,
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }


@router.get("/vulnerability-trend")
def api_vulnerability_trend(
    days: int = Query(30, ge=1, le=365),
    session_id: str = Query(..., min_length=1, max_length=128),
):
    sid = _require_session(session_id)
    return {
        "status": "ok",
        "data": vulnerability_trend(days=days, session_id=sid),
        "period_days": days,
        "session_id": sid,
    }


@router.get("/top-issues")
def api_top_issues(
    limit: int = Query(10, ge=1, le=50),
    session_id: str = Query(..., min_length=1, max_length=128),
):
    sid = _require_session(session_id)
    issues = get_top_issues(limit=limit, session_id=sid)
    return {"status": "ok", "data": issues, "count": len(issues), "session_id": sid}


@router.get("/scan-history")
def api_scan_history(
    days: int = Query(30, ge=1, le=365),
    target: str | None = Query(None, max_length=256),
    limit: int = Query(100, ge=1, le=1000),
    session_id: str = Query(..., min_length=1, max_length=128),
):
    sid = _require_session(session_id)
    scans = get_scan_history(days=days, target=target, limit=limit, session_id=sid)
    return {
        "status": "ok",
        "data": scans,
        "count": len(scans),
        "filter": {"days": days, "target": target, "limit": limit, "session_id": sid},
    }


@router.get("/bundle")
def api_dashboard_bundle(
    days: int = Query(30, ge=1, le=365),
    session_id: str = Query(..., min_length=1, max_length=128),
    history_limit: int = Query(20, ge=1, le=100),
    top_limit: int = Query(10, ge=1, le=50),
):
    """Um request com metrics + trend + top_issues + history."""
    sid = _require_session(session_id)
    data = dashboard_bundle(
        days=days,
        session_id=sid,
        history_limit=history_limit,
        top_limit=top_limit,
    )
    return {
        "status": "ok",
        "session_id": sid,
        "period_days": days,
        **data,
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }


@router.get("/summary")
def api_summary(
    days: int = Query(30, ge=1, le=365),
    session_id: str = Query(..., min_length=1, max_length=128),
):
    sid = _require_session(session_id)
    return {"status": "ok", "summary": summary_report(days=days, session_id=sid)}


@router.delete("/session/{session_id}")
def api_purge_session(session_id: str):
    sid = _require_session(session_id)
    deleted = purge_scans_for_session(sid)
    return {"status": "ok", "session_id": sid, "scans_deleted": deleted}


@router.get("/export")
def api_export(
    format: str = Query("json", pattern="^(json|csv|pdf|xlsx)$"),
    days: int = Query(30, ge=1, le=365),
    session_id: str = Query(..., min_length=1, max_length=128),
):
    sid = _require_session(session_id)
    history = get_scan_history(days=days, limit=10000, session_id=sid)
    metrics = compute_metrics(days=days, session_id=sid)
    fmt = (format or "json").lower()
    date_stamp = datetime.now(timezone.utc).strftime("%Y%m%d")
    file_base = f"darkstar-dashboard-{sid[:8]}-{date_stamp}"

    if fmt == "json":
        return {
            "format": "json",
            "session_id": sid,
            "metrics": metrics,
            "history": history,
            "exported_at": datetime.now(timezone.utc).isoformat(),
        }

    if fmt == "csv":
        output = io.StringIO()
        fields = [
            "target",
            "vulnerability_count",
            "critical",
            "high",
            "medium",
            "low",
            "timestamp",
            "status",
            "scan_type",
            "chat_session_id",
        ]
        writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in history:
            writer.writerow({k: row.get(k) for k in fields})
        data = output.getvalue().encode("utf-8")
        return Response(
            content=data,
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{file_base}.csv"'},
        )

    if fmt == "xlsx":
        trend = vulnerability_trend(days=days, session_id=sid)
        top_issues = get_top_issues(limit=50, session_id=sid)
        data = _build_xlsx(
            sid=sid,
            days=days,
            metrics=metrics,
            history=history,
            trend=trend,
            top_issues=top_issues,
        )
        return Response(
            content=data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{file_base}.xlsx"'},
        )

    if fmt == "pdf":
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas
        except ImportError as exc:
            raise HTTPException(status_code=500, detail="reportlab unavailable") from exc

        buf = io.BytesIO()
        c = canvas.Canvas(buf, pagesize=letter)
        c.setFont("Helvetica-Bold", 16)
        c.drawString(50, 750, "DarkStar — Resumo para o cliente")
        c.setFont("Helvetica", 11)
        y = 720
        lines = [
            f"Conversa: {sid[:24]}",
            f"Periodo: {days} dias",
            f"Varreduras: {metrics.get('total_scans', 0)}",
            f"Criticas (media): {float(metrics.get('avg_critical') or 0):.1f}",
            f"Altas (media): {float(metrics.get('avg_high') or 0):.1f}",
            f"Problemas em aberto: {metrics.get('open_vulnerabilities', 0)}",
            f"Total de achados: {metrics.get('total_vulnerabilities', 0)}",
            "",
            _client_blurb(metrics)[:90],
            "",
            "Alvos recentes:",
        ]
        for line in lines:
            c.drawString(50, y, str(line)[:90])
            y -= 16
        for row in history[:80]:
            line = (
                f"- {row.get('target')} · problemas={row.get('vulnerability_count')} "
                f"crit={row.get('critical')} ({str(row.get('timestamp') or '')[:19]})"
            )
            c.drawString(50, y, line[:95])
            y -= 14
            if y < 60:
                c.showPage()
                y = 750
                c.setFont("Helvetica", 11)
        c.save()
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{file_base}.pdf"'},
        )

    raise HTTPException(status_code=400, detail="Invalid format")
